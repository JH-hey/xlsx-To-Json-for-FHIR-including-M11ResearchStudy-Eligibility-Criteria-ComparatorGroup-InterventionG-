import json
from openpyxl import load_workbook
from collections import OrderedDict
from openpyxl.utils import get_column_letter

def parse_merged_cells(ws):
    """解析合并单元格，返回位置映射字典"""
    merged_values = {}
    for merged_range in ws.merged_cells:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col
        )
        main_value = ws.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_values[(row, col)] = main_value
    return merged_values


def build_nested_structure(ws, merged_values):
    """构建层次化数据结构（修复路径解析）"""
    structure = OrderedDict()

    # 遍历所有列（从第1列开始）
    for col in range(20, ws.max_column + 1):
        path = []
        # 解析前5行表头路径（Excel行号1-5）
        for row in range(1, 6):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value and str(cell_value).strip():
                path.append(str(cell_value).strip())

        # 获取第6行数据值
        value = ws.cell(row=6, column=col).value
        if value is None:
            continue

        # 构建嵌套结构
        current = structure


        for i, key in enumerate(path[:-1]):
            current = current[0] if isinstance(current, list) else current
            if key in current:
                current=current[key]
            else: ##对于coding,modelCharacteristic,attributeEstimate他们得先来个[]，因此我们要分开处理
                if key =="coding":
                    current.setdefault("coding",[])
                    current["coding"].append(OrderedDict())
                    current=current["coding"][0]
                elif key =="modelCharacteristic":
                    current.setdefault("modelCharacteristic", [])
                    current["modelCharacteristic"].append(OrderedDict())
                    current = current["modelCharacteristic"][0]
                elif key =="attributeEstimate":
                    current.setdefault("attributeEstimate", [])
                    current["attributeEstimate"].append(OrderedDict())
                    current = current["attributeEstimate"][0]
                else:
                    current[key] = OrderedDict()
                    current = current[key]

        # 赋值最终键值
        if path:
            last_key = path[-1]
            if isinstance(current, OrderedDict):
                current[last_key] = value if value != "" else None
            else:
                current[0][last_key] = value if value != "" else None

    return structure


def excel_to_json(ws, merged_values):
    data = OrderedDict()

    # 基础字段
    data["resourceType"] = ws['A6'].value
    data["meta"] = OrderedDict()
    data["meta"]["versionId"] = str(ws['B6'].value)
    data["meta"]["lastUpdated"] = str(ws['C6'].value)
    data["meta"]["profile"] = [ws['D6'].value]

    # 固定扩展结构
    data["extension"] = [{
        "extension": [
            {"url": "type", "valueCode": "derived-from"},
            {"url": "quotation", "valueMarkdown": "quote the article text where the information was extracted"},
            {"url": "targetUri", "valueUri": "https://dom-pubs.pericles-prod.literatumonline.com/doi/10.1111/dom.13413"}
        ],
        "url": "http://hl7.org/fhir/uv/ebm/StructureDefinition/relates-to-with-quotation"
    }]

    # 添加上下文
    data["useContext"] = [{
        "code": {
            "system": "https://fevir.net/resources/CodeSystem/179423",
            "code": "evidence-communication",
            "display": "Evidence Communication"
        },
        "valueCodableConcept": {
            "coding": [{
                "system": "https://fevir.net/resources/CodeSystem/179423",
                "code": "InterventionOnlyEvidence",
                "display": "InterventionOnlyEvidence"
            }]
        }
    }]

    # 简单字段
    data["title"] = ws['E6'].value
    data["description"] = ws['F6'].value
    data["assertion"] = ws['G6'].value

    # 解析嵌套结构
    nested = build_nested_structure(ws, merged_values)



    # 处理variableDefinition
    data["variableDefinition"] = []

    for i in range(3):  # 三个对象

        var = OrderedDict()
        start_col=8+4*i
        var["variableRole"] = ws[f'{get_column_letter(start_col)}6'].value
        observed = OrderedDict()
        observed["reference"] = ws[f'{get_column_letter(start_col+1)}6'].value
        observed["type"] = ws[f'{get_column_letter(start_col+2)}6'].value
        observed["display"] = ws[f'{get_column_letter(start_col+3)}6'].value
        var["observed"] = observed
        data["variableDefinition"].append(var)

    # 处理statistic
    data["statistic"] = [nested["statistic"]]


    # 写入文件
    with open("test1.json", "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print("转换完成，结果保存至inter_test0.json")


if __name__ == "__main__":
    wb = load_workbook("InterventionGroupEvidence.xlsx")
    ws = wb.active
    merged_values = parse_merged_cells(ws)

    excel_to_json(ws, merged_values)



