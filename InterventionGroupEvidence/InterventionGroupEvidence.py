import json
from openpyxl import load_workbook
from collections import OrderedDict


def parse_merged_cells(ws):
    """解析合并单元格，返回位置映射字典"""
    merged_values = {}
    for merged_range in ws.merged_cells:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col
        )
        main_value = ws.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_values[(row, col)] = main_value
    return merged_values


def excel_to_json(ws, merged_values):
    data = OrderedDict()

    # 基础字段
    data["resourceType"] = ws['A6'].value
    data["meta"] = OrderedDict()
    data["meta"]["versionId"] = str(ws['B6'].value)
    data["meta"]["lastUpdated"] = str(ws['C6'].value)
    data["meta"]["profile"] = [ws['D6'].value]

    # 固定扩展结构
    data["extension"] = [{
        "extension": [
            {"url": "type", "valueCode": "derived-from"},
            {"url": "quotation", "valueMarkdown": "quote the article text where the information was extracted"},
            {"url": "targetUri", "valueUri": "https://dom-pubs.pericles-prod.literatumonline.com/doi/10.1111/dom.13413"}
        ],
        "url": "http://hl7.org/fhir/uv/ebm/StructureDefinition/relates-to-with-quotation"
    }]

    # 添加上下文
    data["useContext"] = [{
        "code": {
            "system": "https://fevir.net/resources/CodeSystem/179423",
            "code": "evidence-communication",
            "display": "Evidence Communication"
        },
        "valueCodableConcept": {
            "coding": [{
                "system": "https://fevir.net/resources/CodeSystem/179423",
                "code": "InterventionOnlyEvidence",
                "display": "InterventionOnlyEvidence"
            }]
        }
    }]

    # 简单字段
    data["title"] = ws['E6'].value
    data["description"] = ws['F6'].value
    data["assertion"] = ws['G6'].value

    # 解析嵌套结构

# 处理variableDefinition
    data["variableDefinition"] = []
    current_char=None
    for col in range(8,20):
        path=[]
        for row in range(1,6):
            cell_value=merged_values.get((row,col),ws.cell(row=row,column=col).value)
            if cell_value:
                clean_part=str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "variableDefinition" not in path[0]:
            continue
        cell_value = ws.cell(row=6, column=col).value
        value=str(cell_value).strip() if cell_value else ""
        parts=[p for p in path if p not in ['', 'variableDefinition']]
        #切换对象的逻辑
        if len(parts)>=1 and parts[0]=="variableRole":
            if current_char:
                data["variableDefinition"].append(current_char)
                current_char=None
        if len(parts)>=1:
            filed_type=parts[0]
            if not current_char and filed_type=="variableRole":
                current_char=OrderedDict()
                current_char["variableRole"]=value
            elif filed_type=="observed"  and len(parts)>=2:
                ref=current_char.setdefault("observed",{})
                sub_field=parts[1]
                if sub_field=="reference":
                    ref["reference"]=value
                elif sub_field=="type":
                    ref["type"]=value
                elif sub_field=="display":
                    ref["display"]=value
    if current_char:
        data["variableDefinition"].append(current_char)

#处理statistic
    data["statistic"] = []
    current_char = None
    for col in range(20, 50):
        path = []
        for row in range(1, 6):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "statistic" not in path[0]:
            continue
        cell_value = ws.cell(row=6, column=col).value
        value = str(cell_value).strip() if cell_value else ""
        parts = [p for p in path if p not in ['', 'statistic']]
        #处理逻辑特征
        if len(parts) >= 1 and parts[0] == "description":
            if current_char:
                data["statistic"].append(current_char)
                current_char = None
        if len(parts)>=1:
            filed_type=parts[0]
            if not current_char and filed_type=="description":
                current_char=OrderedDict()
                current_char["description"]=value
            elif filed_type=="statisticType":
                if len(parts)>=2:
                    statisticType_field=parts[1]
                    if statisticType_field=="coding" and len(parts)>=3:
                        coding=current_char.setdefault("statisticType",{}).setdefault("coding",[{}])[0]
                        sub_field=parts[2]
                        if sub_field=="system":
                            coding["system"]=value
                        elif sub_field=="code":
                            coding["code"]=value
                        elif sub_field=="display":
                            coding["display"]=value
            elif filed_type=="quantity" and len(parts)>=2:
                Qdict=current_char.setdefault("quantity",{})
                sub_field=parts[1]
                if sub_field=="value":
                    Qdict["value"]=float(value)
                elif sub_field=="unit":
                    Qdict["unit"]=value
            elif filed_type=="attributeEstimate":
                if len(parts)>=2:
                    attribute_dict=current_char.setdefault("attributeEstimate",[{}])[0]
                    sub1_field=parts[1]
                    if sub1_field=="type":
                        type_dict=attribute_dict.setdefault("type",{}).setdefault("coding",[{}])[0]
                        if len(parts)>=4:
                            sub2_field=parts[3]
                            if sub2_field=="system":
                                type_dict["system"]=value
                            elif sub2_field=="code":
                                type_dict["code"]=value
                            elif sub2_field=="display":
                                type_dict["display"]=value
                    elif sub1_field =="level":
                        attribute_dict["level"]=float(value)
                    elif sub1_field=="range":
                        rang_dict=attribute_dict.setdefault("range",{})
                        if len(parts)>=3:
                            sub2_field=parts[2]
                            if sub2_field=="low":
                                rang_dict.setdefault("low",{})
                                sub3_field=parts[3]
                                if sub3_field=="value":
                                    rang_dict["low"]["value"]=float(value)
                                elif sub3_field=="unit":
                                    rang_dict["low"]["unit"]=value
                            elif sub2_field=="high":
                                rang_dict.setdefault("high",{})
                                sub3_field=parts[3]
                                if sub3_field=="value":
                                    rang_dict["high"]["value"]=float(value)
                                elif sub3_field=="unit":
                                    rang_dict["high"]["unit"]=value
            elif filed_type == "modelCharacteristic":
                if len(parts)>=2:
                    if len(parts)>=3:
                        current_char.setdefault("modelCharacteristic",[{}])[0].setdefault("code",{})["text"]=value
    if current_char:
        data["statistic"].append(current_char)



    # 写入文件
    with open("InterventionGroupEvidence.json", "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print("转换完成，结果保存至InterventionGroupEvidence.json")


if __name__ == "__main__":
    wb = load_workbook("InterventionGroupEvidence.xlsx")
    ws = wb.active
    merged_values = parse_merged_cells(ws)
    excel_to_json(ws, merged_values)







