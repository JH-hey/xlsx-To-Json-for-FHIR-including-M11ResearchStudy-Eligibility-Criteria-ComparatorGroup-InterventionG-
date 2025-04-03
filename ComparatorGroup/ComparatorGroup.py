import json
from openpyxl import load_workbook
from collections import OrderedDict

def parse_merged_cells(ws):
    """解析合并单元格，返回位置映射字典"""
    merged_values = {}
    for merged_range in ws.merged_cells:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col
        )
        main_value = ws.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_values[(row, col)] = main_value
    return merged_values

def excel_to_json(ws,merged_values):


    # 创建有序字典保持字段顺序
    data = OrderedDict()
    data["resourceType"] = ws['A5'].value
    data["meta"] = OrderedDict()
    data["meta"]["versionId"] = str(ws['B5'].value)
    data["meta"]["lastUpdated"] = str(ws['C5'].value)
    data["meta"]["profile"] = [ws['D5'].value]

    # 添加固定结构的useContext
    data["useContext"] = [{
        "code": {
            "system": "https://fevir.net/resources/CodeSystem/179423",
            "code": "evidence-communication",
            "display": "Evidence Communication"
        },
        "valueCodableConcept": {
            "coding": [{
                "system": "https://fevir.net/resources/CodeSystem/179423",
                "code": "ComparatorGroup",
                "display": "ComparatorGroup"
            }]
        }
    }]

    # 添加基础字段（根据单元格映射）
    field_mapping = [
        ('title', 'E5'), ('status', 'F5'), ('description', 'G5'),
        ('type', 'H5'), ('membership', 'I5'), ('quantity', 'J5'),
        ('combinationMethod', 'K5')
    ]
    for key, cell in field_mapping:
        value = ws[cell].value
        if key == 'quantity':
            data[key] = int(value)
        else:
            data[key] = str(value).strip() if value else ""

    # 处理characteristic结构
    data["characteristic"] = []
    current_char = None

    # 遍历所有列解析特征数据
    for col in range(12, ws.max_column + 1):
        # 构建列头路径（前4行为路径层级）
        path = []
        for row in range(1, 5):
            cell_value =merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                # 清洗路径组件并转为小写
                clean_part = str(cell_value).strip().lower()
                if clean_part:
                    path.append(clean_part)

        # 跳过非特征列
        if not path or "characteristic" not in path[0]:
            continue

        # 获取数据单元格值（第5行）
        cell_value = ws.cell(row=5, column=col).value
        value = str(cell_value).strip() if cell_value else ""

        # 解析路径组件（示例路径：["characteristic", "1", "code", "text"]）
        parts = [p for p in path if p not in ['', 'characteristic']]

        # 处理特征切换逻辑,切换到下一个对象的Odict。切换的条件是parts[0]==description且current不是空的
        if len(parts) >= 1 and parts[0]=="description":
            # 当出现新的特征序号时，保存当前特征
            if current_char:
                current_char["exclude"] = False
                data["characteristic"].append(current_char)
                current_char = None #一个对象加上去了，就current_char=None，便于后面进入逻辑

        # 处理特征字段
        if len(parts) >= 1:
            field_type = parts[0]#field_type就去获得parts的第一个值，然后就根据filed_type是哪一种，就进入到哪个逻辑

            # 初始化新特征
            if not current_char and field_type == "description": #这里依据是current_char=None，因此就进去了。无论第一次遇到description还是第二次
                current_char = OrderedDict()
                current_char["description"] = value

            # 处理code结构
            elif field_type == "code":
                if len(parts) >= 2:
                    code_field = parts[1]#拿到第二个
                    # code.text处理
                    if code_field == "text":
                        current_char.setdefault("code", {})["text"] = value #先把code的值设为字典，然后取text=value。与reference不同，先判断code和text或code和coding一起写，毕竟中间没有隔阂。因此code是先拿判断再一起写
                    # code.coding处理
                    elif code_field == "coding" and len(parts) >= 3:
                        coding = current_char.setdefault("code", {}).setdefault("coding", [{}])[0]
                        sub_field = parts[2]
                        if sub_field == "system":
                            coding["system"] = value
                        elif sub_field == "code":
                            coding["code"] = value
                        elif sub_field == "display":
                            coding["display"] = value

            # 处理valueReference结构
            elif field_type == "valuereference" and len(parts) >= 2:
                ref = current_char.setdefault("valueReference", {})#default是没有valuereference就写，有就不写。然后取到{}或{...}
                sub_field = parts[1]  #reference是先写再拿进行判断
                if sub_field == "reference":
                    ref["reference"] = value
                elif sub_field == "type":
                    ref["type"] = value
                elif sub_field == "display":
                    ref["display"] = value

    # 添加最后一个特征。等所有列都完事了
    if current_char:
        current_char["exclude"] = False
        data["characteristic"].append(current_char)

    # 写入文件
    output_path = "ComparatorGroup.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"JSON文件已生成：{output_path}")


if __name__ == "__main__":
    wb = load_workbook("ComparatorGroup.xlsx")
    ws = wb.active
    merged_values = parse_merged_cells(ws)
    excel_to_json(ws,merged_values)
















