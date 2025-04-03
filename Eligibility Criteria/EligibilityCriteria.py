import json
from openpyxl import load_workbook
from collections import OrderedDict

def parse_merged_cells(ws):
    """解析合并单元格，返回位置映射字典"""
    merged_values = {}
    for merged_range in ws.merged_cells:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col
        )
        main_value = ws.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_values[(row, col)] = main_value
    return merged_values

def excel_to_json(ws,merged_values):
    data=OrderedDict()
    data["resourceType"]=ws['A6'].value
    data["meta"]=OrderedDict()
    data["meta"]["versionId"]=str(ws['B6'].value)
    data["meta"]["lastUpdated"]=str(ws['C6'].value)
    data["title"]=ws['D6'].value
    data["status"]=ws['E6'].value
    data["description"]=ws['F6'].value
    data["type"]=ws['G6'].value
    data["membership"]=ws['H6'].value
    data["combinationMethod"]=ws['I6'].value
#处理characterisitc部分
    data["characteristic"]=[]
    current_char=None
    for col in range(10,ws.max_column+1):
        path=[]
        for row in range(1,6):
            cell_value=merged_values.get((row,col),ws.cell(row=row,column=col).value)
            if cell_value:
                # 清洗路径组件
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "characteristic" not in path[0]:
            continue
        cell_value = ws.cell(row=6, column=col).value
        value = str(cell_value).strip() if cell_value else ""
        parts = [p for p in path if p not in ['', 'characteristic']]
        #切换对象逻辑
        if len(parts) >= 1 and parts[0]=="description":
            if current_char:
                current_char["exclude"]=False
                data["characteristic"].append(current_char)
                current_char=None

        if len(parts)>=1:
            field_type=parts[0]

            if not current_char and field_type=="description":
                current_char=OrderedDict()
                current_char["description"]=value
            elif field_type =="code":
                if len(parts)>=2:
                    code_field=parts[1]
                    if code_field=="coding" and len(parts)>=3:
                        coding=current_char.setdefault("code",{}).setdefault("coding",[{}])[0]
                        sub_field=parts[2]
                        if sub_field=="system":
                            coding["system"]=value
                        elif sub_field=="code":
                            coding["code"]=value
                        elif sub_field=="display":
                            coding["display"]=value
            elif field_type=="valueRange":
                valueRange_dict=current_char.setdefault("valueRange",{})
                if len(parts)>=2:
                    sub1_field=parts[1]
                    if sub1_field=="low":
                        low_dict=valueRange_dict.setdefault("low",{})
                        sub2_field=parts[2]
                        if sub2_field=="value":
                            low_dict["value"]=int(value)
                        elif sub2_field=="unit":
                            low_dict["unit"]=value
                        elif sub2_field=="system":
                            low_dict["system"]=value
                        elif sub2_field=="code":
                            low_dict["code"]=value
                    elif sub1_field=="high":
                        high_dict=valueRange_dict.setdefault("high",{})
                        sub2_field=parts[2]
                        if sub2_field=="value":
                            high_dict["value"]=int(value)
                        elif sub2_field=="unit":
                            high_dict["unit"]=value
                        elif sub2_field=="system":
                            high_dict["system"]=value
                        elif sub2_field=="code":
                            high_dict["code"]=value
            elif field_type=="valueReference":
                valueReference_dict = current_char.setdefault("valueReference", {})
                if len(parts)>=2:
                    sub_field = parts[1]
                    if sub_field == "reference":
                        valueReference_dict["reference"] = value
                    elif sub_field == "type":
                        valueReference_dict["type"] = value
                    elif sub_field == "display":
                        valueReference_dict["display"] = value
            elif field_type=="valueCodableConcept":
                if len(parts)>=2 and parts[1]=="coding":
                    coding=current_char.setdefault("valueCodableConcept",{}).setdefault("coding",[{}])[0]
                    if len(parts)>=3:
                        sub_field=parts[2]
                        if sub_field=="system":
                            coding["system"]=value
                        elif sub_field=="code":
                            coding["code"]=value
                        elif sub_field=="display":
                            coding["display"]=value

    if current_char:
        current_char["exclude"] = False
        data["characteristic"].append(current_char)





    output_path = "EligibilityCriteria.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"JSON文件已生成：{output_path}")





if __name__ == "__main__":
    wb = load_workbook("EligibilityCriteria.xlsx")
    ws = wb.active
    merged_values = parse_merged_cells(ws)
    excel_to_json(ws,merged_values)
