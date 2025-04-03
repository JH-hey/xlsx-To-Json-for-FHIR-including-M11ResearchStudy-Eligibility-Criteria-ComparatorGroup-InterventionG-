import json
from openpyxl import  load_workbook
from collections import OrderedDict

#解析每个单元格的位置
def parse_merged_cells(ws):

    merged_values={}
    for merged_range in ws.merged_cells:
        min_row,min_col,max_row,max_col=(
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col
        )
        main_value=ws.cell(min_row,min_col).value
        for row in range(min_row,max_row+1):
            for col in range(min_col,max_col+1):
                merged_values[(row,col)]=main_value
    return merged_values


#进行赋值
def excel_to_json(ws,merged_values):
#创建总字典
    data=OrderedDict()
#插入直接定位的数据
    data["resourceType"]=ws['A5'].value.strip()
    data["title"]=ws['B5'].value.strip()
    data["name"]=ws['C5'].value.strip()
    data["status"]=ws['D5'].value.strip()
#处理meta
    data["meta"]=OrderedDict()
    data["meta"]["versionId"]=str(ws['E5'].value).strip()
    data["meta"]["lastUpdated"]=str(ws['F5'].value).strip()
    profile_value=ws['G5'].value.strip()
    if profile_value: #由于profile的值不一定存在，因此这里要判断一下
        data["meta"]["profile"]=[profile_value]
#处理period
    data["period"]=OrderedDict()
    data["period"]["start"]=str(ws['H5'].value).strip()
    data["period"]["end"]=str(ws['I5'].value).strip()
#处理primaryPurposeType
    #采用直接赋值法
    data["primaryPurposeType"]=OrderedDict()
    data["primaryPurposeType"]["coding"]=[]
    Pdict=OrderedDict()
    Pdict["system"]=ws["J5"].value.strip()
    Pdict["code"]=ws["K5"].value.strip()
    Pdict["display"]=ws["L5"].value.strip()
    data["primaryPurposeType"]["coding"].append(Pdict)
#处理recruitment
    data["recruitment"]=OrderedDict()
    data["recruitment"]["targetNumber"]=int(str(ws["M5"].value).strip())
    data["recruitment"]["actualNumber"] =int(str(ws["N5"].value).strip())
    data["recruitment"]["eligibility"]=OrderedDict()
    data["recruitment"]["eligibility"]["reference"]=ws["O5"].value.strip()
    data["recruitment"]["eligibility"]["type"]=ws["P5"].value.strip()
    data["recruitment"]["eligibility"]["display"]=ws["Q5"].value.strip()
#处理phase
    data["phase"]=OrderedDict()
    data["phase"]["coding"]=[]
    Phdict=OrderedDict()
    Phdict["system"]=ws["R5"].value.strip()
    Phdict["code"]=str(ws["S5"].value).strip()
    Phdict["display"]=str(ws["T5"].value).strip()
    data["phase"]["coding"].append(Phdict)


#处理label
    data["label"]=[]
    current_char=None
    #count用来计数的，判断当parts[0]是type的时候，到底是要换一个对象了，还是现在的对象还没写完
    #count应该放在最前面，因为当开始循环列的时候，这个东西就要准备好，一旦开始循环，count就要被一直使用
    count=0
    #遍历该部分所有列，并对每个列进行整理
    for col in range(21,25):
        path=[]
        for row in range(1,5):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                # 清洗路径组件
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "label" not in path[0]:
            continue
        #获取单元格数据（第5行）
        cell_value = ws.cell(row=5, column=col).value
        value = str(cell_value).strip() if cell_value else ""
        #解析路径组间，赋值parts
        parts = [p for p in path if p not in ['', 'label']]

        #判断是否需要切换对象
        if len(parts)>=1 and parts[0]=="type" :
            if count==3:#说明当前current的type的下面system,cod,display已经都走过了，需要换一个对象了。这个是用来判断是什么时候换对象的一个重要标志
                data["label"].append(current_char)
                count=0
                current_char=None
        #判断去哪里
        if len(parts) >=1:
            field_type=parts[0]#拿到parts的第一个值，来判断去哪里
            #当field_type=type时
            if field_type=="type":
                if not current_char:#如果current_char是空的，那么就要另开一个对象。因为current_char是空的，说明上一个对象已经完毕了。如果不是空的，说明这个对象还没有完事，要继续加
                    current_char=OrderedDict()
                type_field=parts[1]#拿到第二个
                if type_field=="coding":
                    coding=current_char.setdefault("type",{}).setdefault("coding",[{}])[0]
                    sub_field= parts[2]
                    if sub_field=="system":
                        coding["system"]=value
                        count+=1
                    elif sub_field=="code":
                        coding["code"]=value
                        count+=1
                    elif sub_field=="display":
                        coding["display"]=value
                        count+=1
            elif field_type=="value":
                current_char["value"]=value

    if current_char:
        data["label"].append(current_char)


#处理identifier
    data["identifier"]=[]
    current_char=None
    for col in range(25,28):
        path=[]
        for row in range(1, 5):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                # 清洗路径组件
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "identifier" not in path[0]:
            continue
        #获取单元格的数据
        cell_value = ws.cell(row=5, column=col).value
        value = str(cell_value).strip() if cell_value else ""
        parts = [p for p in path if p not in ['', 'identifier']]
        #切换对象
        if len(parts)>=1 and parts[0]=="use":
            if current_char:
                data["identifier"].append(current_char)
                current_char=None

        #处理字段
        if len(parts)>=1:
            field_type=parts[0]#拿到第一个，进行分流
            if not current_char and field_type=="use":
                current_char=OrderedDict()
                current_char["use"]=value
            elif field_type=="value":
                current_char["value"]=value
            elif field_type=="assigner":
                if len(parts) >=2:
                    assigner_field = parts[1]
                    if assigner_field=="display":
                        current_char.setdefault("assigner",{})["display"]=value

    if current_char:
        data["identifier"].append(current_char)

# 处理progressStatus
    data["progressStatus"]=[]
    current_char=None
    count=0
    for col in range(28,31):
        path=[]
        for row in range(1, 5):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                # 清洗路径组件
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "progressStatus" not in path[0]:
            continue
        cell_value = ws.cell(row=5, column=col).value
        value=str(cell_value).strip() if cell_value else ""
        parts = [p for p in path if p not in ['', 'progressStatus']]
        #切换对象的逻辑
        if len(parts)>=1 and parts[0]=="state":
            if count==3:
                data["progressStatus"].append(current_char)
                count=0
                current_char=None
        if len(parts)>=1:
            field_type=parts[0]
            if field_type=="state":
                if not current_char:
                    current_char=OrderedDict()
                state_field=parts[1]
                if state_field=="coding":
                    coding=current_char.setdefault("state",{}).setdefault("coding",[{}])[0]
                    sub_field=parts[2]
                    if sub_field=="system":
                        coding["system"]=value
                        count+=1
                    elif sub_field=="code":
                        coding["code"]=value
                        count+=1
                    elif sub_field=="display":
                        coding["display"]=value
                        count+=1
    if current_char:
        data["progressStatus"].append(current_char)

#处理condition
    data["condition"]=[]
    current_char=None
    for col in range(31,33):
        path=[]
        for row in range(1,5):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                # 清洗路径组件
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "condition" not in path[0]:
            continue
        cell_value = ws.cell(row=5, column=col).value
        value=str(cell_value).strip() if cell_value else ""
        parts=[p for p in path if p not in ['', 'condition']]
        if len(parts) >= 1 and parts[0]=="text":
            if current_char:
                data["condition"].append(current_char)
                current_char=None
        if len(parts) >= 1:
            field_type=parts[0]
            if not current_char and field_type=="text":
                current_char=OrderedDict()
                current_char["text"]=value
    if current_char:
        data["condition"].append(current_char)

#处理region:
    data["region"]=[]
    current_char=None
    for col in range(33,34):
        path=[]
        for row in range(1,5):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                # 清洗路径组件
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "region" not in path[0]:
            continue
        cell_value = ws.cell(row=5, column=col).value
        value=str(cell_value).strip() if cell_value else ""
        parts=[p for p in path if p not in ['', 'region']]
        if len(parts)>=1 and parts[0]=="text":
            if current_char:
                data["region"].append(current_char)
                current_char=None
        if len(parts)>=1:
            field_type=parts[0]
            if not current_char and field_type=="text":
                current_char=OrderedDict()
                current_char["text"]=value
    if current_char:
        data["region"].append(current_char)

#处理studyDesign:
    data["studyDesign"]=[]
    current_char=None
    count=0
    for col in range(34,46):
        path=[]
        for row in range(1,5):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                # 清洗路径组件
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "studyDesign" not in path[0]:
            continue
        cell_value = ws.cell(row=5, column=col).value
        value=str(cell_value).strip() if cell_value else ""
        parts=[p for p in path if p not in ['', 'studyDesign']]
        if len(parts)>=1 and parts[0]=="coding":
            if count==3:
                data["studyDesign"].append(current_char)
                count=0
                current_char=None

        if len(parts)>=1:
            field_type=parts[0]
            if field_type=="coding":
                if not current_char:
                    current_char=OrderedDict()
                coding=current_char.setdefault("coding",[{}])[0]
                sub_field=parts[1]
                if sub_field=="system":
                    coding["system"]=value
                    count+=1
                elif sub_field=="code":
                    coding["code"]=value
                    count+=1
                elif sub_field=="display":
                    coding["display"]=value
                    count+=1
    if current_char:
        data["studyDesign"].append(current_char)


#处理relatesTo:
    data["relatesTo"]=[]
    current_char=None
    for col in range(46,50):
        path=[]
        for row in range(1,5):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                # 清洗路径组件
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "relatesTo" not in path[0]:
            continue
        cell_value = ws.cell(row=5, column=col).value
        value=str(cell_value).strip() if cell_value else ""
        parts=[p for p in path if p not in ['', 'relatesTo']]
        if len(parts)>=1 and parts[0]=="type":
            if current_char:
                data["relatesTo"].append(current_char)
                current_char=None
        if len(parts)>=1:
            field_type=parts[0]
            if not current_char and field_type=="type":
                current_char=OrderedDict()
                current_char["type"]=value
            elif field_type =="targetCanonical":
                current_char["targetCanonical"]=value
            elif field_type =="targetMarkdown":
                current_char["targetMarkdown"]=value
    if current_char:
        data["relatesTo"].append(current_char)

#associatedParty
    data["associatedParty"]=[]
    current_char=None
    for col in range(50,64):
        path=[]
        for row in range(1,5):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                # 清洗路径组件
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "associatedParty" not in path[0]:
            continue
        cell_value = ws.cell(row=5, column=col).value
        value=str(cell_value).strip() if cell_value else ""
        parts=[p for p in path if p not in ['', 'associatedParty']]
        if len(parts)>=1 and parts[0]=="name":
            if current_char:
                data["associatedParty"].append(current_char)
                current_char=None
        if len(parts)>=1:
            field_type=parts[0]
            if not current_char and field_type=="name":
                current_char=OrderedDict()
                current_char["name"]=value
            elif field_type=="role":
                if len(parts)>=2:
                    role_field=parts[1]
                    if role_field=="coding" and len(parts)>=3:
                        coding=current_char.setdefault("role", {}).setdefault("coding", [{}])[0]
                        sub_field=parts[2]
                        if sub_field=="system":
                            coding["system"]=value
                        elif sub_field=="code":
                            coding["code"]=value
                        elif sub_field=="display":
                            coding["display"]=value
            elif field_type=="classifier":
                if len(parts)>=2:
                    classifier_field=parts[1]
                    if classifier_field=="coding" and len(parts)>=3:
                        #这里的classfier本身也是一个数组，因为原json的classifier就一个对象，因此我们这里就把对象一个一个append，就直接设定就一个对象，即[{}]
                        coding=current_char.setdefault("classifier", [{}])[0].setdefault("coding", [{}])[0]
                        sub_field=parts[2]
                        if sub_field=="system":
                            coding["system"]=value
                        elif sub_field=="code":
                            coding["code"]=value
                        elif sub_field=="display":
                            coding["display"]=value
    if current_char:
        data["associatedParty"].append(current_char)

#处理comparisonGroup
    data["comparisonGroup"]=[]
    current_char=None
    for col in range(64,86):
        path=[]
        for row in range(1,5):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                # 清洗路径组件
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "comparisonGroup" not in path[0]:
            continue
        cell_value = ws.cell(row=5, column=col).value
        value=str(cell_value).strip() if cell_value else ""
        parts=[p for p in path if p not in ['', 'comparisonGroup']]
        if len(parts)>=1 and parts[0]=="name":
            if current_char:
                data["comparisonGroup"].append(current_char)
                current_char=None
        if len(parts)>=1:
            field_type=parts[0]
            if not current_char and field_type=="name":
                current_char=OrderedDict()
                current_char["name"]=value
            elif field_type=="type":
                if len(parts)>=2:
                    type_field=parts[1]
                    if type_field=="coding" and len(parts)>=3:
                        coding=current_char.setdefault("type", {}).setdefault("coding", [{}])[0]
                        sub_field=parts[2]
                        if sub_field=="system":
                            coding["system"]=value
                        elif sub_field=="code":
                            coding["code"]=value
                        elif sub_field=="display":
                            coding["display"]=value
            elif field_type=="description":
                current_char["description"]=value
            elif field_type=="intendedExposure" and len(parts)>=2:
                ref=current_char.setdefault("intendedExposure", [{}])[0]
                sub_field=parts[1]
                if sub_field=="reference":
                    ref["reference"]=value
                if sub_field=="type":
                    ref["type"]=value
                if sub_field=="display":
                    ref["display"]=value
            elif field_type=="observedGroup" and len(parts)>=2:
                ref=current_char.setdefault("observedGroup", [{}])[0]
                sub_field=parts[1]
                if sub_field=="reference":
                    ref["reference"]=value
                if sub_field=="type":
                    ref["type"]=value
                if sub_field=="display":
                    ref["display"]=value
    if current_char:
        data["comparisonGroup"].append(current_char)
#处理outcomeMeasure:
    data["outcomeMeasure"]=[]
    current_char=None
    for col in range(86,94):
        path=[]
        for row in range(1,5):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                # 清洗路径组件
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or "outcomeMeasure" not in path[0]:
            continue
        cell_value = ws.cell(row=5, column=col).value
        value=str(cell_value).strip() if cell_value else ""
        parts=[p for p in path if p not in ['', 'outcomeMeasure']]
        #切换对象的逻辑
        if len(parts)>=1 and parts[0]=="name":
            if current_char:
                data["outcomeMeasure"].append(current_char)
                current_char=None
        if len(parts)>=1:
            field_type=parts[0]
            if not current_char and field_type=="name":
                current_char=OrderedDict()
                current_char["name"]=value
            elif field_type=="type":
                if len(parts)>=2:
                    type_field=parts[1]
                    if type_field=="coding" and len(parts)>=3:
                        #这里的type本身也是一个数组，因为原json的type就一个对象，因此我们这里就把对象一个一个append，就直接设定就一个对象，即[{}]
                        coding=current_char.setdefault("type", [{}])[0].setdefault("coding", [{}])[0]
                        sub_field=parts[2]
                        if sub_field=="system":
                            coding["system"]=value
                        elif sub_field=="code":
                            coding["code"]=value
                        elif sub_field=="display":
                            coding["display"]=value
            elif field_type=="description":
                current_char["description"]=value
            elif field_type=="reference":
                if len(parts)>=2:
                    ref=current_char.setdefault("reference", {})
                    sub_field=parts[1]
                    if sub_field=="reference":
                        ref["reference"]=value
                    elif sub_field=="type":
                        ref["type"]=value
                    elif sub_field=="display":
                        ref["display"]=value
    if current_char:
        data["outcomeMeasure"].append(current_char)

#处理result
    data["result"]=[]
    current_char=None
    for col in range(94,115):
        path=[]
        for row in range(1,5):
            cell_value = merged_values.get((row, col), ws.cell(row=row, column=col).value)
            if cell_value:
                # 清洗路径组件
                clean_part = str(cell_value).strip()
                if clean_part:
                    path.append(clean_part)
        if not path or  "result" not in path[0]:
            continue
        cell_value = ws.cell(row=5, column=col).value
        value=str(cell_value).strip() if cell_value else ""
        parts=[p for p in path if p not in ['', 'result']]
        if len(parts)>=1 and parts[0]=="reference":
            if current_char:
                data["result"].append(current_char)
                current_char=None

        if len(parts)>=1:
            field_type=parts[0]
            if not current_char and field_type=="reference":
                current_char=OrderedDict()
                current_char["reference"]=value
            elif field_type=="type":
                current_char["type"]=value
            elif field_type=="display":
                current_char["display"]=value
    if current_char:
        data["result"].append(current_char)


#添加固定结构的extension
    data["extension"]=[
    {
      "extension": [
        {
          "url": "type",
          "valueCode": "derived-from"
        },
        {
          "url": "quotation",
          "valueMarkdown": "quote the article text where the information was extracted"
        },
        {
          "url": "targetUri",
          "valueUri": "https://dom-pubs.pericles-prod.literatumonline.com/doi/10.1111/dom.13413"
        }
      ],
      "url": "http://hl7.org/fhir/uv/ebm/StructureDefinition/relates-to-with-quotation"
    }
  ]


    output_path = "M11ResearchStudy.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"JSON文件已生成：{output_path}")



if __name__ == '__main__':
    wb=load_workbook("M11ResearchStudy.xlsx")
    ws=wb.active
    merged_values=parse_merged_cells(ws)
    excel_to_json(ws,merged_values)

